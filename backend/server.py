from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException, Form
from fastapi.responses import JSONResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime
import openpyxl
import xlrd
import io
import base64


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")


# Define Models
class StatusCheck(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    client_name: str
    timestamp: datetime = Field(default_factory=datetime.utcnow)

class StatusCheckCreate(BaseModel):
    client_name: str

class WorkItem(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    work_no: str
    work_description: str
    estimated_cost: Optional[float] = None
    completion_time: Optional[str] = None
    location: Optional[str] = None
    category: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)

class TenderNotice(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tender_no: str
    notice_title: str
    organization: Optional[str] = None
    publication_date: Optional[datetime] = None
    last_date_submission: Optional[datetime] = None
    work_items: List[WorkItem] = []
    excel_file_name: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)
    status: str = "active"  # active, closed, cancelled

class BidderProfile(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_name: str
    contact_person: str
    email: str
    phone: str
    address: str
    registration_no: Optional[str] = None
    pan_no: Optional[str] = None
    gst_no: Optional[str] = None
    experience_years: Optional[int] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)

class BidSubmission(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tender_id: str
    work_item_id: str
    bidder_id: str
    quoted_amount: float
    completion_time_proposed: Optional[str] = None
    technical_documents: Optional[List[str]] = []  # Base64 encoded files
    financial_documents: Optional[List[str]] = []  # Base64 encoded files
    remarks: Optional[str] = None
    submitted_at: datetime = Field(default_factory=datetime.utcnow)
    status: str = "submitted"  # submitted, under_review, accepted, rejected

# Create Models for API requests
class TenderNoticeCreate(BaseModel):
    tender_no: str
    notice_title: str
    organization: Optional[str] = None
    publication_date: Optional[datetime] = None
    last_date_submission: Optional[datetime] = None

class BidderProfileCreate(BaseModel):
    company_name: str
    contact_person: str
    email: str
    phone: str
    address: str
    registration_no: Optional[str] = None
    pan_no: Optional[str] = None
    gst_no: Optional[str] = None
    experience_years: Optional[int] = None

class BidSubmissionCreate(BaseModel):
    tender_id: str
    work_item_id: str
    bidder_id: str
    quoted_amount: float
    completion_time_proposed: Optional[str] = None
    remarks: Optional[str] = None


# Utility functions for Excel processing
def parse_excel_file(file_content: bytes, filename: str) -> List[WorkItem]:
    """Parse Excel file and extract work items"""
    work_items = []
    
    try:
        if filename.endswith('.xlsx'):
            # Handle .xlsx files
            workbook = openpyxl.load_workbook(io.BytesIO(file_content))
            sheet = workbook.active
            
            # Assuming the Excel has headers in the first row
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value)
            
            # Process data rows
            for row_num in range(2, sheet.max_row + 1):
                row_data = {}
                for col_num, header in enumerate(headers, 1):
                    cell_value = sheet.cell(row=row_num, column=col_num).value
                    if header:
                        row_data[header.lower().strip()] = cell_value
                
                # Create WorkItem from row data
                if row_data.get('work_no') or row_data.get('work_description'):
                    work_item = WorkItem(
                        work_no=str(row_data.get('work_no', f'WORK_{row_num-1}')),
                        work_description=str(row_data.get('work_description', '')),
                        estimated_cost=float(row_data.get('estimated_cost', 0)) if row_data.get('estimated_cost') else None,
                        completion_time=str(row_data.get('completion_time', '')) if row_data.get('completion_time') else None,
                        location=str(row_data.get('location', '')) if row_data.get('location') else None,
                        category=str(row_data.get('category', '')) if row_data.get('category') else None
                    )
                    work_items.append(work_item)
                    
        elif filename.endswith('.xls'):
            # Handle .xls files
            workbook = xlrd.open_workbook(file_contents=file_content)
            sheet = workbook.sheet_by_index(0)
            
            # Get headers from first row
            headers = []
            for col in range(sheet.ncols):
                headers.append(sheet.cell_value(0, col))
            
            # Process data rows
            for row_num in range(1, sheet.nrows):
                row_data = {}
                for col_num, header in enumerate(headers):
                    cell_value = sheet.cell_value(row_num, col_num)
                    if header:
                        row_data[header.lower().strip()] = cell_value
                
                # Create WorkItem from row data
                if row_data.get('work_no') or row_data.get('work_description'):
                    work_item = WorkItem(
                        work_no=str(row_data.get('work_no', f'WORK_{row_num}')),
                        work_description=str(row_data.get('work_description', '')),
                        estimated_cost=float(row_data.get('estimated_cost', 0)) if row_data.get('estimated_cost') else None,
                        completion_time=str(row_data.get('completion_time', '')) if row_data.get('completion_time') else None,
                        location=str(row_data.get('location', '')) if row_data.get('location') else None,
                        category=str(row_data.get('category', '')) if row_data.get('category') else None
                    )
                    work_items.append(work_item)
                    
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error parsing Excel file: {str(e)}")
    
    return work_items


# API Endpoints
@api_router.get("/")
async def root():
    return {"message": "Tender Management System API"}

@api_router.post("/status", response_model=StatusCheck)
async def create_status_check(input: StatusCheckCreate):
    status_dict = input.dict()
    status_obj = StatusCheck(**status_dict)
    _ = await db.status_checks.insert_one(status_obj.dict())
    return status_obj

@api_router.get("/status", response_model=List[StatusCheck])
async def get_status_checks():
    status_checks = await db.status_checks.find().to_list(1000)
    return [StatusCheck(**status_check) for status_check in status_checks]

# Tender Notice Management
@api_router.post("/tender-notices/upload-excel")
async def upload_tender_excel(
    file: UploadFile = File(...),
    tender_no: str = Form(...),
    notice_title: str = Form(...),
    organization: str = Form(None),
    publication_date: str = Form(None),
    last_date_submission: str = Form(None)
):
    """Upload Excel file with tender notice and work items"""
    
    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")
    
    try:
        # Read file content
        file_content = await file.read()
        
        # Parse Excel file to extract work items
        work_items = parse_excel_file(file_content, file.filename)
        
        # Create tender notice
        tender_notice = TenderNotice(
            tender_no=tender_no,
            notice_title=notice_title,
            organization=organization,
            publication_date=datetime.fromisoformat(publication_date) if publication_date else None,
            last_date_submission=datetime.fromisoformat(last_date_submission) if last_date_submission else None,
            work_items=work_items,
            excel_file_name=file.filename
        )
        
        # Save to database
        result = await db.tender_notices.insert_one(tender_notice.dict())
        
        return {
            "message": "Tender notice uploaded successfully",
            "tender_id": tender_notice.id,
            "work_items_count": len(work_items),
            "work_items": [item.dict() for item in work_items]
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")

@api_router.get("/tender-notices", response_model=List[TenderNotice])
async def get_tender_notices():
    """Get all tender notices"""
    tender_notices = await db.tender_notices.find().to_list(1000)
    return [TenderNotice(**notice) for notice in tender_notices]

@api_router.get("/tender-notices/{tender_id}", response_model=TenderNotice)
async def get_tender_notice(tender_id: str):
    """Get specific tender notice by ID"""
    tender_notice = await db.tender_notices.find_one({"id": tender_id})
    if not tender_notice:
        raise HTTPException(status_code=404, detail="Tender notice not found")
    return TenderNotice(**tender_notice)

@api_router.delete("/tender-notices/{tender_id}")
async def delete_tender_notice(tender_id: str):
    """Delete tender notice"""
    result = await db.tender_notices.delete_one({"id": tender_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Tender notice not found")
    return {"message": "Tender notice deleted successfully"}

# Bidder Management
@api_router.post("/bidders", response_model=BidderProfile)
async def create_bidder_profile(bidder: BidderProfileCreate):
    """Create new bidder profile"""
    try:
        bidder_profile = BidderProfile(**bidder.dict())
        await db.bidder_profiles.insert_one(bidder_profile.dict())
        return bidder_profile
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error creating bidder profile: {str(e)}")

@api_router.get("/bidders", response_model=List[BidderProfile])
async def get_bidder_profiles():
    """Get all bidder profiles"""
    bidders = await db.bidder_profiles.find().to_list(1000)
    return [BidderProfile(**bidder) for bidder in bidders]

@api_router.get("/bidders/{bidder_id}", response_model=BidderProfile)
async def get_bidder_profile(bidder_id: str):
    """Get specific bidder profile"""
    bidder = await db.bidder_profiles.find_one({"id": bidder_id})
    if not bidder:
        raise HTTPException(status_code=404, detail="Bidder not found")
    return BidderProfile(**bidder)

@api_router.put("/bidders/{bidder_id}", response_model=BidderProfile)
async def update_bidder_profile(bidder_id: str, bidder: BidderProfileCreate):
    """Update bidder profile"""
    try:
        updated_bidder = bidder.dict()
        result = await db.bidder_profiles.update_one(
            {"id": bidder_id},
            {"$set": updated_bidder}
        )
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Bidder not found")
        
        updated_bidder_doc = await db.bidder_profiles.find_one({"id": bidder_id})
        return BidderProfile(**updated_bidder_doc)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error updating bidder profile: {str(e)}")

# Bid Submission Management
@api_router.post("/bids", response_model=BidSubmission)
async def submit_bid(bid: BidSubmissionCreate):
    """Submit a bid for a work item"""
    
    try:
        # Verify tender and work item exist
        tender_notice = await db.tender_notices.find_one({"id": bid.tender_id})
        if not tender_notice:
            raise HTTPException(status_code=404, detail="Tender notice not found")
        
        # Check if work item exists in the tender
        work_item_exists = any(item['id'] == bid.work_item_id for item in tender_notice.get('work_items', []))
        if not work_item_exists:
            raise HTTPException(status_code=404, detail="Work item not found in this tender")
        
        # Verify bidder exists
        bidder = await db.bidder_profiles.find_one({"id": bid.bidder_id})
        if not bidder:
            raise HTTPException(status_code=404, detail="Bidder not found")
        
        bid_submission = BidSubmission(**bid.dict())
        await db.bid_submissions.insert_one(bid_submission.dict())
        return bid_submission
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error submitting bid: {str(e)}")

@api_router.get("/bids/tender/{tender_id}")
async def get_bids_for_tender(tender_id: str):
    """Get all bids for a specific tender"""
    bids = await db.bid_submissions.find({"tender_id": tender_id}).to_list(1000)
    return [BidSubmission(**bid) for bid in bids]

@api_router.get("/bids/work-item/{work_item_id}")
async def get_bids_for_work_item(work_item_id: str):
    """Get all bids for a specific work item"""
    bids = await db.bid_submissions.find({"work_item_id": work_item_id}).to_list(1000)
    return [BidSubmission(**bid) for bid in bids]

@api_router.get("/bids/bidder/{bidder_id}")
async def get_bids_by_bidder(bidder_id: str):
    """Get all bids submitted by a specific bidder"""
    bids = await db.bid_submissions.find({"bidder_id": bidder_id}).to_list(1000)
    return [BidSubmission(**bid) for bid in bids]

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
