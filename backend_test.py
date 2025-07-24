#!/usr/bin/env python3
"""
Backend API Testing for Tender Management System
Tests the newly added tender management API endpoints to ensure Excel upload and bidder data bugs are fixed.
"""

import requests
import json
import io
import openpyxl
from datetime import datetime
import time

# Backend URL from frontend/.env
BACKEND_URL = "https://26d32f48-af6c-460f-a81a-e953e8cce2df.preview.emergentagent.com/api"

class TenderAPITester:
    def __init__(self):
        self.base_url = BACKEND_URL
        self.session = requests.Session()
        self.test_results = []
        
    def log_result(self, test_name, success, message, response_data=None):
        """Log test results"""
        result = {
            "test": test_name,
            "success": success,
            "message": message,
            "timestamp": datetime.now().isoformat(),
            "response_data": response_data
        }
        self.test_results.append(result)
        status = "✅ PASS" if success else "❌ FAIL"
        print(f"{status} {test_name}: {message}")
        if response_data and not success:
            print(f"   Response: {response_data}")
    
    def create_test_excel_file(self):
        """Create a test Excel file with work items"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        # Headers
        headers = ["work_no", "work_description", "estimated_cost", "completion_time", "location", "category"]
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)
        
        # Sample data
        test_data = [
            ["WORK001", "Road Construction Phase 1", 500000, "6 months", "Mumbai", "Infrastructure"],
            ["WORK002", "Bridge Repair and Maintenance", 250000, "3 months", "Delhi", "Maintenance"],
            ["WORK003", "Water Pipeline Installation", 750000, "8 months", "Bangalore", "Utilities"]
        ]
        
        for row_idx, row_data in enumerate(test_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.getvalue()
    
    def test_api_root(self):
        """Test API root endpoint"""
        try:
            response = self.session.get(f"{self.base_url}/")
            if response.status_code == 200:
                data = response.json()
                self.log_result("API Root", True, "API is accessible", data)
                return True
            else:
                self.log_result("API Root", False, f"HTTP {response.status_code}: {response.text}")
                return False
        except Exception as e:
            self.log_result("API Root", False, f"Connection error: {str(e)}")
            return False
    
    def test_excel_upload(self):
        """Test POST /api/tender-notices/upload-excel"""
        try:
            # Create test Excel file
            excel_data = self.create_test_excel_file()
            
            # Prepare form data
            files = {
                'file': ('test_tender.xlsx', excel_data, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            }
            
            form_data = {
                'tender_no': 'TND2025001',
                'notice_title': 'Municipal Infrastructure Development Project',
                'organization': 'Mumbai Municipal Corporation',
                'publication_date': '2025-01-15T10:00:00',
                'last_date_submission': '2025-02-15T17:00:00'
            }
            
            response = self.session.post(
                f"{self.base_url}/tender-notices/upload-excel",
                files=files,
                data=form_data
            )
            
            if response.status_code == 200:
                data = response.json()
                if 'tender_id' in data and 'work_items_count' in data:
                    self.log_result("Excel Upload", True, f"Excel uploaded successfully. Tender ID: {data['tender_id']}, Work items: {data['work_items_count']}", data)
                    return data['tender_id']  # Return tender ID for later tests
                else:
                    self.log_result("Excel Upload", False, "Missing expected fields in response", data)
                    return None
            else:
                self.log_result("Excel Upload", False, f"HTTP {response.status_code}: {response.text}")
                return None
                
        except Exception as e:
            self.log_result("Excel Upload", False, f"Error: {str(e)}")
            return None
    
    def test_create_bidder(self):
        """Test POST /api/bidders"""
        try:
            bidder_data = {
                "company_name": "ABC Construction Pvt Ltd",
                "contact_person": "Rajesh Kumar",
                "email": "rajesh@abcconstruction.com",
                "phone": "+91-9876543210",
                "address": "123 Industrial Area, Sector 15, Gurgaon, Haryana 122001",
                "registration_no": "REG2024001",
                "pan_no": "ABCDE1234F",
                "gst_no": "07ABCDE1234F1Z5",
                "experience_years": 15
            }
            
            response = self.session.post(
                f"{self.base_url}/bidders",
                json=bidder_data,
                headers={'Content-Type': 'application/json'}
            )
            
            if response.status_code == 200:
                data = response.json()
                if 'id' in data and data['company_name'] == bidder_data['company_name']:
                    self.log_result("Create Bidder", True, f"Bidder created successfully. ID: {data['id']}", data)
                    return data['id']  # Return bidder ID for later tests
                else:
                    self.log_result("Create Bidder", False, "Missing expected fields in response", data)
                    return None
            else:
                self.log_result("Create Bidder", False, f"HTTP {response.status_code}: {response.text}")
                return None
                
        except Exception as e:
            self.log_result("Create Bidder", False, f"Error: {str(e)}")
            return None
    
    def test_get_tender_notices(self):
        """Test GET /api/tender-notices"""
        try:
            response = self.session.get(f"{self.base_url}/tender-notices")
            
            if response.status_code == 200:
                data = response.json()
                if isinstance(data, list):
                    self.log_result("Get Tender Notices", True, f"Retrieved {len(data)} tender notices", {"count": len(data)})
                    return data
                else:
                    self.log_result("Get Tender Notices", False, "Response is not a list", data)
                    return None
            else:
                self.log_result("Get Tender Notices", False, f"HTTP {response.status_code}: {response.text}")
                return None
                
        except Exception as e:
            self.log_result("Get Tender Notices", False, f"Error: {str(e)}")
            return None
    
    def test_get_bidders(self):
        """Test GET /api/bidders"""
        try:
            response = self.session.get(f"{self.base_url}/bidders")
            
            if response.status_code == 200:
                data = response.json()
                if isinstance(data, list):
                    self.log_result("Get Bidders", True, f"Retrieved {len(data)} bidders", {"count": len(data)})
                    return data
                else:
                    self.log_result("Get Bidders", False, "Response is not a list", data)
                    return None
            else:
                self.log_result("Get Bidders", False, f"HTTP {response.status_code}: {response.text}")
                return None
                
        except Exception as e:
            self.log_result("Get Bidders", False, f"Error: {str(e)}")
            return None
    
    def test_submit_bid(self, tender_id, bidder_id, work_item_id):
        """Test POST /api/bids"""
        if not all([tender_id, bidder_id, work_item_id]):
            self.log_result("Submit Bid", False, "Missing required IDs for bid submission")
            return None
            
        try:
            bid_data = {
                "tender_id": tender_id,
                "work_item_id": work_item_id,
                "bidder_id": bidder_id,
                "quoted_amount": 450000.0,
                "completion_time_proposed": "5 months",
                "remarks": "We have extensive experience in similar projects and can deliver within timeline."
            }
            
            response = self.session.post(
                f"{self.base_url}/bids",
                json=bid_data,
                headers={'Content-Type': 'application/json'}
            )
            
            if response.status_code == 200:
                data = response.json()
                if 'id' in data and data['tender_id'] == tender_id:
                    self.log_result("Submit Bid", True, f"Bid submitted successfully. ID: {data['id']}", data)
                    return data['id']
                else:
                    self.log_result("Submit Bid", False, "Missing expected fields in response", data)
                    return None
            else:
                self.log_result("Submit Bid", False, f"HTTP {response.status_code}: {response.text}")
                return None
                
        except Exception as e:
            self.log_result("Submit Bid", False, f"Error: {str(e)}")
            return None
    
    def test_error_handling(self):
        """Test error handling scenarios"""
        print("\n=== Testing Error Handling ===")
        
        # Test invalid Excel upload
        try:
            files = {'file': ('test.txt', b'not an excel file', 'text/plain')}
            form_data = {'tender_no': 'TEST', 'notice_title': 'Test'}
            
            response = self.session.post(
                f"{self.base_url}/tender-notices/upload-excel",
                files=files,
                data=form_data
            )
            
            if response.status_code == 400:
                self.log_result("Invalid File Upload", True, "Correctly rejected non-Excel file")
            else:
                self.log_result("Invalid File Upload", False, f"Expected 400, got {response.status_code}")
        except Exception as e:
            self.log_result("Invalid File Upload", False, f"Error: {str(e)}")
        
        # Test invalid bidder data
        try:
            invalid_bidder = {"company_name": ""}  # Missing required fields
            
            response = self.session.post(
                f"{self.base_url}/bidders",
                json=invalid_bidder,
                headers={'Content-Type': 'application/json'}
            )
            
            if response.status_code in [400, 422]:
                self.log_result("Invalid Bidder Data", True, "Correctly rejected invalid bidder data")
            else:
                self.log_result("Invalid Bidder Data", False, f"Expected 400/422, got {response.status_code}")
        except Exception as e:
            self.log_result("Invalid Bidder Data", False, f"Error: {str(e)}")
    
    def run_all_tests(self):
        """Run all tests in sequence"""
        print("=== Starting Tender Management API Tests ===")
        print(f"Backend URL: {self.base_url}")
        print()
        
        # Test API connectivity
        if not self.test_api_root():
            print("❌ API is not accessible. Stopping tests.")
            return
        
        print("\n=== Core Functionality Tests ===")
        
        # Test Excel upload
        tender_id = self.test_excel_upload()
        
        # Test bidder creation
        bidder_id = self.test_create_bidder()
        
        # Test data retrieval
        tender_notices = self.test_get_tender_notices()
        bidders = self.test_get_bidders()
        
        # Test bid submission (if we have all required IDs)
        work_item_id = None
        if tender_notices and len(tender_notices) > 0:
            # Get work item ID from the first tender
            for notice in tender_notices:
                if notice.get('work_items') and len(notice['work_items']) > 0:
                    work_item_id = notice['work_items'][0]['id']
                    break
        
        if tender_id and bidder_id and work_item_id:
            self.test_submit_bid(tender_id, bidder_id, work_item_id)
        else:
            self.log_result("Submit Bid", False, "Could not test bid submission - missing required data")
        
        # Test error handling
        self.test_error_handling()
        
        # Print summary
        self.print_summary()
    
    def print_summary(self):
        """Print test summary"""
        print("\n" + "="*60)
        print("TEST SUMMARY")
        print("="*60)
        
        passed = sum(1 for result in self.test_results if result['success'])
        total = len(self.test_results)
        
        print(f"Total Tests: {total}")
        print(f"Passed: {passed}")
        print(f"Failed: {total - passed}")
        print(f"Success Rate: {(passed/total)*100:.1f}%")
        
        print("\nDetailed Results:")
        for result in self.test_results:
            status = "✅" if result['success'] else "❌"
            print(f"{status} {result['test']}: {result['message']}")
        
        # Critical issues
        critical_failures = [r for r in self.test_results if not r['success'] and r['test'] in ['Excel Upload', 'Create Bidder', 'Submit Bid']]
        if critical_failures:
            print(f"\n⚠️  CRITICAL ISSUES FOUND ({len(critical_failures)}):")
            for failure in critical_failures:
                print(f"   - {failure['test']}: {failure['message']}")

if __name__ == "__main__":
    tester = TenderAPITester()
    tester.run_all_tests()